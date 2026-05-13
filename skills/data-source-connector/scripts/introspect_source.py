#!/usr/bin/env python3
"""
introspect_source.py — Source-agnostic data-model introspection.

Reads a source descriptor (JSON on stdin or via --source-file) and emits a
normalized `data-model.json` to stdout (or --output).

Supported source types out of the box:
  - excel   (requires: openpyxl)
  - csv     (requires: pandas)
  - parquet (requires: pandas + pyarrow)

Other source types (databricks, snowflake, sqlserver, etc.) emit a stub
data-model.json with `openQuestions[]` asking the orchestrator to provide
the missing connection details or to use a different ingestion path.

Usage:
    python introspect_source.py --source-file source.json --output data-model.json
    cat source.json | python introspect_source.py > data-model.json

Example source.json (Excel):
    {
      "type": "excel",
      "connection": { "path": "C:/data/sales.xlsx" },
      "options": { "sampleRows": 5 }
    }
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

TMDL_SAFE_RE = re.compile(r"[^A-Za-z0-9_]+")


def to_tmdl_name(name: str) -> str:
    """Sanitize a physical name into a TMDL-safe identifier."""
    cleaned = TMDL_SAFE_RE.sub("_", name.strip()).strip("_")
    if not cleaned:
        cleaned = "unnamed"
    if cleaned[0].isdigit():
        cleaned = "_" + cleaned
    return cleaned


def infer_type(values: list[Any]) -> tuple[str, str]:
    """Infer (dataType, sourceProviderType) from sample values."""
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "string", "nvarchar(65535)"

    def _is_int(v: Any) -> bool:
        if isinstance(v, bool):
            return False
        if isinstance(v, int):
            return True
        if isinstance(v, str):
            try:
                int(v)
                return True
            except ValueError:
                return False
        return False

    def _is_float(v: Any) -> bool:
        if isinstance(v, bool):
            return False
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            try:
                float(v)
                return True
            except ValueError:
                return False
        return False

    def _is_bool(v: Any) -> bool:
        if isinstance(v, bool):
            return True
        if isinstance(v, str):
            return v.strip().lower() in {"true", "false"}
        return False

    def _is_datetime(v: Any) -> bool:
        if isinstance(v, datetime):
            return True
        if isinstance(v, str):
            for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                        "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    datetime.strptime(v, fmt)
                    return True
                except ValueError:
                    continue
        return False

    if all(_is_bool(v) for v in non_null):
        return "boolean", "bit"
    if all(_is_datetime(v) for v in non_null):
        return "dateTime", "datetime2"
    if all(_is_int(v) for v in non_null):
        return "int64", "bigint"
    if all(_is_float(v) for v in non_null):
        return "double", "double"
    return "string", "nvarchar(65535)"


def infer_role(table_name: str, columns: list[dict]) -> str:
    """Heuristically classify a table."""
    n = table_name.lower()
    if n.startswith("fact_") or n in {"orders", "sales", "transactions", "events"}:
        return "fact"
    if n.startswith("dim_") or n in {"customers", "products", "geography"}:
        return "dimension"
    if "date" in n or "calendar" in n:
        return "date dimension"

    has_dates = any(c["dataType"] == "dateTime" for c in columns)
    has_doubles = any(c["dataType"] == "double" for c in columns)
    has_fk_like = sum(1 for c in columns if c["name"].endswith(("_key", "_id")))

    if has_doubles and has_fk_like >= 2:
        return "fact"
    if has_dates and len(columns) > 5:
        return "date dimension"
    return "dimension"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# Adapters
# ---------------------------------------------------------------------------

def introspect_excel(conn: dict, opts: dict) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for Excel introspection. "
            "Install it: pip install openpyxl"
        ) from e

    path = conn.get("path")
    if not path or not Path(path).exists():
        return _stub(
            "excel", conn,
            open_questions=[{
                "id": "q1",
                "scope": "connection",
                "question": f"Excel file not found at `{path}`. Please provide the correct path."
            }]
        )

    sample_rows = int(opts.get("sampleRows", 5))
    wb = load_workbook(path, data_only=True, read_only=True)
    tables: list[dict] = []
    m_templates: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

        data_rows: list[tuple] = []
        for i, row in enumerate(rows_iter):
            if i >= 100:
                break
            data_rows.append(row)

        columns_by_idx: list[list] = [[] for _ in header]
        for row in data_rows:
            for i, val in enumerate(row):
                if i < len(header):
                    columns_by_idx[i].append(val)

        cols: list[dict] = []
        for i, h in enumerate(header):
            dtype, ptype = infer_type(columns_by_idx[i])
            tmdl_name = to_tmdl_name(h)
            col: dict = {
                "name": tmdl_name,
                "physicalName": h,
                "dataType": dtype,
                "sourceProviderType": ptype,
                "nullable": True,
            }
            if i == 0 and dtype in ("int64", "string"):
                col["isPrimaryKey"] = True
            if tmdl_name.endswith(("_key", "_id")) and i != 0:
                col["isForeignKey"] = True
            cols.append(col)

        sample = []
        for row in data_rows[:sample_rows]:
            sample.append({
                cols[i]["name"]: (row[i].isoformat() if isinstance(row[i], datetime) else row[i])
                for i in range(min(len(row), len(cols)))
            })

        table_name = to_tmdl_name(sheet_name)
        role = infer_role(table_name, cols)
        tables.append({
            "name": table_name,
            "physicalName": sheet_name,
            "role": role,
            "rowCountEstimate": ws.max_row - 1 if ws.max_row else 0,
            "columns": cols,
            "sampleRows": sample,
        })
        m_templates[table_name] = (
            'let\n'
            f'    Source = Excel.Workbook(File.Contents("{path}"), null, true),\n'
            f'    {to_tmdl_name(sheet_name)}_Sheet = Source{{[Item="{sheet_name}",Kind="Sheet"]}}[Data],\n'
            f'    #"Promoted Headers" = Table.PromoteHeaders({to_tmdl_name(sheet_name)}_Sheet, [PromoteAllScalars=true])\n'
            'in\n'
            '    #"Promoted Headers"'
        )

    relationships = _infer_relationships(tables)
    open_questions = _open_questions_for_relationships(relationships)

    return {
        "source": {"type": "excel", "connection": {"path": path}, "discoveredAt": now_iso()},
        "tables": tables,
        "relationships": relationships,
        "mCodeAdapter": {"mode": "import", "templates": m_templates},
        "openQuestions": open_questions,
    }


def introspect_csv(conn: dict, opts: dict) -> dict:
    try:
        import pandas as pd
    except ImportError as e:
        raise RuntimeError(
            "pandas is required for CSV introspection. Install it: pip install pandas"
        ) from e

    path = conn.get("path")
    if not path or not Path(path).exists():
        return _stub(
            "csv", conn,
            open_questions=[{
                "id": "q1",
                "scope": "connection",
                "question": f"CSV path not found at `{path}`."
            }]
        )

    sample_rows = int(opts.get("sampleRows", 5))
    paths = [Path(path)] if Path(path).is_file() else sorted(Path(path).glob("*.csv"))
    tables: list[dict] = []
    m_templates: dict[str, str] = {}

    for p in paths:
        df = pd.read_csv(p, nrows=200)
        cols: list[dict] = []
        for i, col in enumerate(df.columns):
            values = df[col].dropna().tolist()
            dtype, ptype = infer_type(values)
            tmdl_name = to_tmdl_name(str(col))
            entry: dict = {
                "name": tmdl_name,
                "physicalName": str(col),
                "dataType": dtype,
                "sourceProviderType": ptype,
                "nullable": bool(df[col].isna().any()),
            }
            if i == 0:
                entry["isPrimaryKey"] = True
            if tmdl_name.endswith(("_key", "_id")) and i != 0:
                entry["isForeignKey"] = True
            cols.append(entry)

        table_name = to_tmdl_name(p.stem)
        role = infer_role(table_name, cols)
        sample = df.head(sample_rows).to_dict(orient="records")
        tables.append({
            "name": table_name,
            "physicalName": p.name,
            "role": role,
            "rowCountEstimate": int(len(df)),
            "columns": cols,
            "sampleRows": sample,
        })
        m_templates[table_name] = (
            'let\n'
            f'    Source = Csv.Document(File.Contents("{p.as_posix()}"), [Delimiter=",", Columns={len(cols)}, Encoding=65001, QuoteStyle=QuoteStyle.Csv]),\n'
            '    #"Promoted Headers" = Table.PromoteHeaders(Source, [PromoteAllScalars=true])\n'
            'in\n'
            '    #"Promoted Headers"'
        )

    relationships = _infer_relationships(tables)
    open_questions = _open_questions_for_relationships(relationships)

    return {
        "source": {"type": "csv", "connection": {"path": str(path)}, "discoveredAt": now_iso()},
        "tables": tables,
        "relationships": relationships,
        "mCodeAdapter": {"mode": "import", "templates": m_templates},
        "openQuestions": open_questions,
    }


def introspect_parquet(conn: dict, opts: dict) -> dict:
    try:
        import pandas as pd
    except ImportError as e:
        raise RuntimeError(
            "pandas + pyarrow are required for Parquet introspection."
        ) from e
    path = conn.get("path")
    if not path or not Path(path).exists():
        return _stub("parquet", conn, open_questions=[{
            "id": "q1", "scope": "connection",
            "question": f"Parquet path not found at `{path}`."
        }])
    df = pd.read_parquet(path)
    cols = [
        {
            "name": to_tmdl_name(c),
            "physicalName": c,
            "dataType": infer_type(df[c].dropna().tolist())[0],
            "sourceProviderType": infer_type(df[c].dropna().tolist())[1],
            "nullable": bool(df[c].isna().any()),
        }
        for c in df.columns
    ]
    table_name = to_tmdl_name(Path(path).stem)
    table = {
        "name": table_name,
        "physicalName": Path(path).name,
        "role": infer_role(table_name, cols),
        "rowCountEstimate": int(len(df)),
        "columns": cols,
        "sampleRows": df.head(int(opts.get("sampleRows", 5))).to_dict(orient="records"),
    }
    return {
        "source": {"type": "parquet", "connection": {"path": str(path)}, "discoveredAt": now_iso()},
        "tables": [table],
        "relationships": [],
        "mCodeAdapter": {
            "mode": "import",
            "templates": {table_name: f'let\n    Source = Parquet.Document(File.Contents("{path}"))\nin\n    Source'},
        },
        "openQuestions": [],
    }


def introspect_unsupported(source_type: str, conn: dict) -> dict:
    """For sources that require live credentials, emit a stub with open questions."""
    return _stub(source_type, conn, open_questions=[{
        "id": "q1",
        "scope": "connection",
        "question": (
            f"Live introspection for `{source_type}` requires credentials and a "
            f"native driver. Please either (a) provide tables/columns manually as a "
            f"data-model.json patch, or (b) export the relevant tables to CSV/Excel "
            f"so this script can introspect them."
        ),
    }])


def _stub(source_type: str, conn: dict, open_questions: list[dict]) -> dict:
    return {
        "source": {"type": source_type, "connection": conn, "discoveredAt": now_iso()},
        "tables": [],
        "relationships": [],
        "mCodeAdapter": {"mode": "import", "templates": {}},
        "openQuestions": open_questions,
    }


# ---------------------------------------------------------------------------
# Relationship inference
# ---------------------------------------------------------------------------

def _infer_relationships(tables: list[dict]) -> list[dict]:
    rels: list[dict] = []
    table_by_name = {t["name"]: t for t in tables}
    pk_index: dict[str, tuple[str, str]] = {}
    for t in tables:
        for c in t["columns"]:
            if c.get("isPrimaryKey"):
                pk_index[c["name"]] = (t["name"], c["name"])

    for t in tables:
        for c in t["columns"]:
            if not c.get("isForeignKey"):
                continue
            cn = c["name"]
            if cn in pk_index and pk_index[cn][0] != t["name"]:
                target_table, target_col = pk_index[cn]
                rels.append({
                    "from": {"table": t["name"], "column": cn},
                    "to": {"table": target_table, "column": target_col},
                    "cardinality": "many-to-one",
                    "isActive": True,
                    "inferredFrom": "naming convention",
                })
                continue
            base = cn[:-4] if cn.endswith("_key") else cn[:-3] if cn.endswith("_id") else cn
            for cand_name, cand in table_by_name.items():
                if cand_name == t["name"]:
                    continue
                if base in cand_name or cand_name in base:
                    for cc in cand["columns"]:
                        if cc.get("isPrimaryKey"):
                            rels.append({
                                "from": {"table": t["name"], "column": cn},
                                "to": {"table": cand_name, "column": cc["name"]},
                                "cardinality": "many-to-one",
                                "isActive": True,
                                "inferredFrom": "naming convention",
                            })
                            break
                    break
    return rels


def _open_questions_for_relationships(rels: list[dict]) -> list[dict]:
    out = []
    for i, r in enumerate(rels, start=1):
        if r.get("inferredFrom") == "naming convention":
            out.append({
                "id": f"r{i}",
                "scope": "relationship",
                "question": (
                    f"I inferred a relationship `{r['from']['table']}.{r['from']['column']}` → "
                    f"`{r['to']['table']}.{r['to']['column']}` from naming. Is that correct?"
                ),
            })
    return out


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

ADAPTERS = {
    "excel": introspect_excel,
    "csv": introspect_csv,
    "parquet": introspect_parquet,
}

LIVE_CREDS_SOURCES = {
    "databricks", "snowflake", "bigquery", "synapse", "sqlserver",
    "postgres", "mysql", "oracle", "odata", "rest", "sharepoint",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Introspect a data source and emit data-model.json")
    parser.add_argument("--source-file", help="Path to source descriptor JSON")
    parser.add_argument("--output", help="Path to write data-model.json (default: stdout)")
    args = parser.parse_args()

    if args.source_file:
        with open(args.source_file, "r", encoding="utf-8") as f:
            descriptor = json.load(f)
    else:
        descriptor = json.load(sys.stdin)

    stype = descriptor.get("type", "").lower()
    conn = descriptor.get("connection", {})
    opts = descriptor.get("options", {})

    if stype in ADAPTERS:
        result = ADAPTERS[stype](conn, opts)
    elif stype in LIVE_CREDS_SOURCES:
        result = introspect_unsupported(stype, conn)
    else:
        print(f"Unknown source type: {stype!r}", file=sys.stderr)
        return 2

    out = json.dumps(result, indent=2, default=str)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(out)
    else:
        print(out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
